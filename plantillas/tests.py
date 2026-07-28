from io import BytesIO
from datetime import datetime

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import PlantillaGenerada
from .views import construir_recomendaciones_analisis


class RecomendacionesAnalisisTests(TestCase):
    def test_destaca_carga_desigual_y_rechazos(self):
        analisis_usuarios = [
            {
                "usuario__username": "Ana",
                "gestiones": 20,
                "tasa_rechazo": 10,
                "rechazados": 2,
            },
            {
                "usuario__username": "Luis",
                "gestiones": 5,
                "tasa_rechazo": 30,
                "rechazados": 3,
            },
        ]

        recomendaciones = construir_recomendaciones_analisis(
            analisis_usuarios=analisis_usuarios,
            promedio_gestiones_por_usuario=12.5,
            total_gestiones=25,
            total_duplicados=3,
        )

        titulos = [item["titulo"] for item in recomendaciones]

        self.assertIn("Carga desigual entre usuarios", titulos)
        self.assertIn("Revisión de rechazos", titulos)
        self.assertIn("Seguimiento de duplicados", titulos)


class VerificarCedulaTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="analista",
            password="prueba-segura",
        )
        self.client.force_login(self.user)

    def crear_registro(self, cedula, gestion):
        return PlantillaGenerada.objects.create(
            usuario=self.user,
            gestion=gestion,
            cedula=cedula,
            nombre_cliente="CLIENTE PRUEBA",
            nombre_plantilla="PRUEBA",
            resultado="PROCEDE",
            distribuidor="DTS PRUEBA",
            respuesta="Respuesta",
        )

    def test_busca_la_cedula_completa_y_devuelve_historial(self):
        self.crear_registro("123456789", "GESTION-1")

        response = self.client.get(
            reverse("verificar_cedula"),
            {"cedula": "123456789"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["existe"])
        self.assertEqual(response.json()["total"], 1)

    def test_no_hace_coincidencias_parciales(self):
        self.crear_registro("123456789", "GESTION-1")

        response = self.client.get(
            reverse("verificar_cedula"),
            {"cedula": "3456"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()["existe"])

    def test_encuentra_cedulas_con_distintos_separadores(self):
        for indice, separador in enumerate(("/", ",", "-", " ")):
            primera = f"C0228995{indice}"
            segunda = f"15583675973{indice}"
            self.crear_registro(
                f"{primera}{separador}{segunda}",
                f"GESTION-{indice}",
            )

            for cedula in (primera.lower(), segunda):
                with self.subTest(
                    separador=repr(separador),
                    cedula=cedula,
                ):
                    response = self.client.get(
                        reverse("verificar_cedula"),
                        {"cedula": cedula},
                    )

                    self.assertEqual(response.status_code, 200)
                    self.assertTrue(response.json()["existe"])
                    self.assertEqual(
                        response.json()["historial"][0]["gestion"],
                        f"GESTION-{indice}",
                    )

    def test_cedula_vacia_no_consulta_todo_el_historial(self):
        self.crear_registro("", "GESTION-1")

        response = self.client.get(reverse("verificar_cedula"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()["existe"])

    def test_exporta_excel_valido_con_filtros(self):
        self.crear_registro("111111111", "GESTION-INCLUIDA")
        self.crear_registro("222222222", "GESTION-EXCLUIDA")

        response = self.client.get(
            reverse("exportar_excel"),
            {"q": "GESTION-INCLUIDA"},
        )
        contenido = b"".join(response.streaming_content)
        libro = load_workbook(BytesIO(contenido), read_only=True)
        hoja = libro["Historial"]
        filas = list(hoja.iter_rows(values_only=True))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(filas[0][2], "Gestión")
        self.assertEqual(len(filas), 2)
        self.assertEqual(filas[1][2], "GESTION-INCLUIDA")


class ReportesOperativosTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="analista-reportes",
            password="prueba-segura",
        )
        self.client.force_login(self.user)

    def crear_registro(self, gestion, resultado, fecha, plantilla="PLANTILLA A"):
        registro = PlantillaGenerada.objects.create(
            usuario=self.user,
            gestion=gestion,
            cedula=f"CED-{gestion}",
            nombre_cliente="CLIENTE",
            nombre_plantilla=plantilla,
            resultado=resultado,
            distribuidor="DTS PRUEBA",
            respuesta="Respuesta",
        )
        PlantillaGenerada.objects.filter(id=registro.id).update(
            fecha=timezone.make_aware(fecha)
        )
        return registro

    def test_compara_con_periodo_anterior_y_analiza_plantillas(self):
        self.crear_registro(
            "ACTUAL-1",
            "PROCEDE",
            datetime(2026, 7, 10, 9, 0),
        )
        self.crear_registro(
            "ACTUAL-2",
            "RECHAZO",
            datetime(2026, 7, 11, 14, 0),
        )
        self.crear_registro(
            "ANTERIOR-1",
            "PROCEDE",
            datetime(2026, 7, 9, 10, 0),
        )

        response = self.client.get(
            reverse("reportes"),
            {
                "fecha_inicio": "2026-07-10",
                "fecha_fin": "2026-07-11",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_gestiones"], 2)
        self.assertEqual(response.context["comparacion_periodo"]["total"], 1)
        self.assertEqual(
            response.context["comparacion_periodo"]["variacion_total"],
            100.0,
        )
        self.assertEqual(response.context["plantillas"][0]["rechazados"], 1)
        self.assertEqual(response.context["plantillas"][0]["tasa_rechazo"], 50.0)

    def test_filtro_horario_se_aplica_a_metricas(self):
        self.crear_registro(
            "DENTRO-HORARIO",
            "PROCEDE",
            datetime(2026, 7, 10, 9, 0),
        )
        self.crear_registro(
            "FUERA-HORARIO",
            "PROCEDE",
            datetime(2026, 7, 10, 19, 0),
        )

        response = self.client.get(
            reverse("reportes"),
            {
                "fecha_inicio": "2026-07-10",
                "fecha_fin": "2026-07-10",
                "hora_inicio": "08:00",
                "hora_fin": "17:00",
            },
        )

        self.assertEqual(response.context["total_gestiones"], 1)
        self.assertEqual(
            sum(item["total"] for item in response.context["actividad_horas"]),
            1,
        )
