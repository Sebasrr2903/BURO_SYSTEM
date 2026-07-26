from django.shortcuts import get_object_or_404, render, redirect
from django.http import FileResponse, HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta, timezone
from django.utils import timezone
from urllib3 import request
from .models import PlantillaGenerada
from django.db.models.functions import TruncDate
from django.contrib.auth.models import User
import json
import tempfile

from .models import PlantillaGenerada


def construir_recomendaciones_analisis(analisis_usuarios, promedio_gestiones_por_usuario, total_gestiones, total_duplicados):
    recomendaciones = []

    if not analisis_usuarios:
        return recomendaciones

    usuario_mas_activo = max(
        analisis_usuarios,
        key=lambda item: (item.get("gestiones") or 0, item.get("usuario__username") or "")
    )

    if usuario_mas_activo:
        recomendaciones.append({
            "titulo": "Usuario con más actividad",
            "descripcion": (
                f"{usuario_mas_activo.get('usuario__username') or 'Sin usuario'} concentra "
                f"{usuario_mas_activo.get('gestiones', 0)} gestiones en este período."
            ),
        })

    usuarios_sobre_promedio = [
        u for u in analisis_usuarios
        if (u.get("gestiones") or 0) > (promedio_gestiones_por_usuario or 0)
    ]

    if len(analisis_usuarios) > 1 and usuarios_sobre_promedio:
        usuario_destacado = usuarios_sobre_promedio[0]
        recomendaciones.append({
            "titulo": "Carga desigual entre usuarios",
            "descripcion": (
                f"{usuario_destacado.get('usuario__username') or 'Un usuario'} está por encima del promedio "
                f"de gestiones y podría generar sobrecarga; conviene revisar la distribución."
            ),
        })

    usuario_mas_rechazos = None
    if analisis_usuarios:
        usuario_mas_rechazos = max(
            analisis_usuarios,
            key=lambda item: (item.get("tasa_rechazo", 0), item.get("rechazados", 0))
        )

    if usuario_mas_rechazos and (usuario_mas_rechazos.get("gestiones") or 0) and (usuario_mas_rechazos.get("tasa_rechazo", 0) or 0) >= 15:
        recomendaciones.append({
            "titulo": "Revisión de rechazos",
            "descripcion": (
                f"{usuario_mas_rechazos.get('usuario__username') or 'Sin usuario'} registra "
                f"{usuario_mas_rechazos.get('tasa_rechazo', 0)}% de rechazos; revisar calidad y proceso ayudaría a reducirlos."
            ),
        })

    if total_duplicados:
        recomendaciones.append({
            "titulo": "Seguimiento de duplicados",
            "descripcion": (
                f"Hay {total_duplicados} cédulas con múltiples gestiones en este filtro; revisar si corresponden "
                "a reintentos o a casos que necesitan mejor seguimiento."
            ),
        })

    if total_gestiones and len(analisis_usuarios) > 1:
        recomendaciones.append({
            "titulo": "Seguimiento de desempeño",
            "descripcion": "Monitorea semanalmente a los usuarios con mayor volumen y a los que muestran más rechazos para mejorar la productividad.",
        })

    return recomendaciones


@login_required
def inicio(request):
    return render(request, 'index.html')



@login_required
def verificar_cedula(request):

    cedula = request.GET.get('cedula', '').strip().upper()

    if not cedula:
        return JsonResponse({"existe": False})

    historial = (
        PlantillaGenerada.objects
        .filter(
            Q(cedula=cedula)
            | Q(cedula_busqueda_1=cedula)
            | Q(cedula_busqueda_2=cedula)
        )
        .select_related("usuario")
        .order_by("-fecha")
    )
    registros = list(historial[:10])

    if registros:
        registro = registros[0]

        return JsonResponse({
            "existe": True,
            "fecha": timezone.localtime(
                registro.fecha
            ).strftime("%d/%m/%Y %H:%M"),
            "usuario": registro.usuario.username if registro.usuario else "",
            "cedula": registro.cedula,
            "nombre_cliente": registro.nombre_cliente,
            "resultado": registro.resultado,
            "distribuidor": registro.distribuidor,
            "respuesta": registro.respuesta,
            "total": historial.count(),

            "historial": [
                {
                    "fecha": timezone.localtime(
                        r.fecha
                    ).strftime("%d/%m/%Y %H:%M"),
                    "gestion": r.gestion,
                    "usuario": r.usuario.username if r.usuario else "",
                    "resultado": r.resultado,
                    "distribuidor": r.distribuidor,
                    "respuesta": r.respuesta
                }
                for r in registros
            ]
        })

    return JsonResponse({
        "existe": False
    })


@csrf_exempt
def guardar_plantilla(request):



    if request.method == "POST":

        data = json.loads(request.body)

        print("ENTRE A GUARDAR")
        print(data)
    
        
        PlantillaGenerada.objects.create(
            usuario=request.user,
            gestion=data.get("gestion"),
            distribuidor=data.get("distribuidor"),
            cedula=data.get("cedula"),
            nombre_cliente=data.get("nombre_cliente"),
            nombre_plantilla=data.get("nombre_plantilla"),
            resultado=data.get("resultado"),
            respuesta=data.get("respuesta")
        )

        return JsonResponse({"success": True})

    return JsonResponse({"success": False})


from django.db.models import Q, Count, Max, F
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from .models import Preset


@login_required
def historial(request):
    registros = PlantillaGenerada.objects.all().order_by('-fecha')

    busqueda = request.GET.get('q')
    usuario = request.GET.get('usuario')
    distribuidor = request.GET.get('distribuidor')
    resultado = request.GET.get('resultado')

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if busqueda:
        registros = registros.filter(
            Q(gestion__icontains=busqueda) |
            Q(nombre_cliente__icontains=busqueda) |
            Q(cedula__icontains=busqueda)
        )

    if distribuidor:
        registros = registros.filter(
            distribuidor__icontains=distribuidor
        )

    if usuario:
        registros = registros.filter(
            usuario__username=usuario
        )

    if resultado == "RECHAZADOS":
        registros = registros.exclude(
            resultado__in=["PROCEDE", "NO PROCEDE"]
        )
    elif resultado:
        registros = registros.filter(resultado=resultado)

    if fecha_inicio:
        registros = registros.filter(fecha__date__gte=fecha_inicio)

    if fecha_fin:
        registros = registros.filter(fecha__date__lte=fecha_fin)

    total = registros.count()

    procede = registros.filter(resultado="PROCEDE").count()
    no_procede = registros.filter(resultado="NO PROCEDE").count()
    rechazados = registros.exclude(resultado__in=["PROCEDE", "NO PROCEDE"]).count()

    # PAGINACIÓN
    limite = request.GET.get('limite', 10)
    paginator = Paginator(registros, int(limite))
    page_number = request.GET.get('page')
    registros = paginator.get_page(page_number)

    usuarios = User.objects.order_by('username')

    return render(
        request,
        'historial.html',
        {
            'registros': registros,
            'distribuidor': distribuidor,
            'resultado': resultado,
            'total': total,
            'procede': procede,
            'no_procede': no_procede,
            'rechazados': rechazados,
            'busqueda': busqueda,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'usuarios': usuarios,
            'usuario_seleccionado': usuario,
        }
    )


@login_required
def presets_list(request):
    presets = Preset.objects.filter(user=request.user).order_by('-creado_en')
    data = [
        {
            'id': p.id,
            'nombre': p.nombre,
            'fecha_inicio': p.fecha_inicio.isoformat() if p.fecha_inicio else '',
            'fecha_fin': p.fecha_fin.isoformat() if p.fecha_fin else '',
            'hora_inicio': p.hora_inicio.strftime('%H:%M') if p.hora_inicio else '',
            'hora_fin': p.hora_fin.strftime('%H:%M') if p.hora_fin else '',
        }
        for p in presets
    ]
    return JsonResponse({'presets': data})


@login_required
@require_POST
def presets_create(request):
    nombre = request.POST.get('nombre')
    fecha_inicio = request.POST.get('fecha_inicio') or None
    fecha_fin = request.POST.get('fecha_fin') or None
    hora_inicio = request.POST.get('hora_inicio') or None
    hora_fin = request.POST.get('hora_fin') or None

    p = Preset.objects.create(
        user=request.user,
        nombre=nombre,
        fecha_inicio=fecha_inicio or None,
        fecha_fin=fecha_fin or None,
        hora_inicio=hora_inicio or None,
        hora_fin=hora_fin or None,
    )

    return JsonResponse({'success': True, 'id': p.id})


@login_required
@require_POST
def presets_delete(request, preset_id):
    p = get_object_or_404(Preset, id=preset_id, user=request.user)
    p.delete()
    return JsonResponse({'success': True})
    

    

    total = registros.count()

    procede = registros.filter(
        resultado="PROCEDE"
    ).count()

    no_procede = registros.filter(
        resultado="NO PROCEDE"
    ).count()

    rechazados = registros.exclude(
    resultado__in=["PROCEDE", "NO PROCEDE"]
    ).count()

    # PAGINACIÓN AL FINAL
    limite = request.GET.get('limite', 10)

    paginator = Paginator(
        registros,
        int(limite)
    )

    page_number = request.GET.get('page')

    registros = paginator.get_page(
        page_number
    )
    usuarios = User.objects.order_by('username')

    return render(
        request,
        'historial.html',
        {
            'registros': registros,
            'distribuidor': distribuidor,
            'resultado': resultado,
            'total': total,
            'procede': procede,
            'no_procede': no_procede,
            'rechazados': rechazados,
            'busqueda': busqueda,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'usuarios': usuarios,
            'usuario_seleccionado': usuario,
        }
    )

@login_required
def exportar_excel(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    hora_inicio = request.GET.get("hora_inicio")
    hora_fin = request.GET.get("hora_fin")
    usuario = request.GET.get("usuario")
    resultado = request.GET.get("resultado")
    distribuidor = request.GET.get("distribuidor")
    busqueda = request.GET.get("q", "").strip()

    registros = PlantillaGenerada.objects.all()

    try:
        if fecha_inicio:
            inicio_dt = datetime.strptime(
                f"{fecha_inicio} {hora_inicio or '00:00'}",
                "%Y-%m-%d %H:%M",
            )
            registros = registros.filter(
                fecha__gte=timezone.make_aware(inicio_dt)
            )

        if fecha_fin:
            if hora_fin:
                fin_dt = datetime.strptime(
                    f"{fecha_fin} {hora_fin}",
                    "%Y-%m-%d %H:%M",
                )
                registros = registros.filter(
                    fecha__lte=timezone.make_aware(fin_dt)
                )
            else:
                fin_exclusivo = (
                    datetime.strptime(fecha_fin, "%Y-%m-%d")
                    + timedelta(days=1)
                )
                registros = registros.filter(
                    fecha__lt=timezone.make_aware(fin_exclusivo)
                )
    except (TypeError, ValueError):
        pass

    if usuario:
        registros = registros.filter(usuario__username=usuario)
    if distribuidor:
        registros = registros.filter(distribuidor=distribuidor)
    if resultado == "RECHAZADOS":
        registros = registros.exclude(
            resultado__in=["PROCEDE", "NO PROCEDE"]
        )
    elif resultado:
        registros = registros.filter(resultado=resultado)
    if busqueda:
        registros = registros.filter(
            Q(gestion__icontains=busqueda)
            | Q(cedula__icontains=busqueda)
            | Q(nombre_cliente__icontains=busqueda)
            | Q(distribuidor__icontains=busqueda)
            | Q(usuario__username__icontains=busqueda)
        )

    filas = (
        registros
        .order_by("-fecha")
        .values_list(
            "fecha",
            "usuario__username",
            "gestion",
            "nombre_cliente",
            "cedula",
            "resultado",
            "distribuidor",
            "respuesta",
        )
        .iterator(chunk_size=2000)
    )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Historial")
    ws.freeze_panes = "A2"
    encabezados = [
        "Fecha",
        "Usuario",
        "Gestión",
        "Cliente",
        "Cédula",
        "Resultado",
        "Distribuidor",
        "Respuesta",
    ]
    fila_encabezado = []
    for titulo in encabezados:
        celda = WriteOnlyCell(ws, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0D6EFD")
        celda.alignment = Alignment(horizontal="center")
        fila_encabezado.append(celda)
    ws.append(fila_encabezado)

    for indice, ancho in enumerate(
        [18, 20, 18, 34, 24, 18, 32, 80],
        start=1,
    ):
        ws.column_dimensions[get_column_letter(indice)].width = ancho

    def texto_excel(valor):
        texto = ILLEGAL_CHARACTERS_RE.sub("", str(valor or ""))
        if texto.startswith(("=", "+", "-", "@")):
            texto = f"'{texto}"
        return texto[:32767]

    total_filas = 1
    for (
        fecha,
        username,
        gestion,
        nombre_cliente,
        cedula,
        estado,
        dts,
        respuesta,
    ) in filas:
        total_filas += 1
        ws.append([
            timezone.localtime(fecha).strftime("%d/%m/%Y %H:%M"),
            texto_excel(username),
            texto_excel(gestion),
            texto_excel(nombre_cliente),
            texto_excel(cedula),
            texto_excel(estado),
            texto_excel(dts),
            texto_excel(respuesta),
        ])

    ws.auto_filter.ref = f"A1:H{total_filas}"

    archivo = tempfile.SpooledTemporaryFile(
        max_size=20 * 1024 * 1024,
        mode="w+b",
    )
    wb.save(archivo)
    archivo.seek(0)

    fecha_archivo = timezone.localdate().strftime("%Y%m%d")
    return FileResponse(
        archivo,
        as_attachment=True,
        filename=f"historial_{fecha_archivo}.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )



@login_required
def limpiar_historial(request):

    if not request.user.groups.filter(
        name='Admin'
    ).exists():

        return redirect('/')

    PlantillaGenerada.objects.all().delete()

    return redirect('/historial/')

@login_required
def ultima_gestion(request):

    registro = PlantillaGenerada.objects.filter(
        usuario=request.user
    ).order_by('-fecha').first()

    if not registro:
        return JsonResponse({
            "success": False
        })

    datos = {
        "success": True,
        "distribuidor": registro.distribuidor,
        "gestion": registro.gestion,
        "cedula": registro.cedula,
        "nombre_cliente": registro.nombre_cliente,
        "plantilla": registro.nombre_plantilla,
        "resultado": registro.resultado,
        "respuesta": registro.respuesta,
    }

    registro.delete()

    return JsonResponse(datos)


@login_required
def limpiar_por_fecha(request):

    if not request.user.groups.filter(
        name='Admin'
    ).exists():

        return redirect('historial')

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio and fecha_fin:

        PlantillaGenerada.objects.filter(
            fecha__date__range=[
                fecha_inicio,
                fecha_fin
            ]
        ).delete()

    return redirect('historial')

@login_required
def eliminar_usuario(request, user_id):

    if not request.user.groups.filter(
        name='Admin'
    ).exists():

        return redirect('/')

    usuario = get_object_or_404(
        User,
        id=user_id
    )

    # Evitar borrarse a sí mismo
    if usuario == request.user:
        return redirect('/crear-usuario/')

    usuario.delete()

    return redirect('/crear-usuario/')





from .models import PlantillaGenerada


@login_required
def reportes(request):

    # =====================================================
    # FILTROS
    # =====================================================

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    usuario = request.GET.get("usuario")
    resultado = request.GET.get("resultado")
    distribuidor = request.GET.get("distribuidor")
    hora_inicio = request.GET.get("hora_inicio")
    hora_fin = request.GET.get("hora_fin")
    busqueda = request.GET.get("q", "").strip()
    periodo = request.GET.get("periodo", "30_dias")

    # La carga inicial usa un período acotado para no procesar todo el
    # histórico. Los filtros permiten ampliar el rango cuando se necesite.
    if not fecha_inicio and not fecha_fin and periodo != "todo":
        hoy = timezone.localdate()
        fecha_fin = hoy.isoformat()
        fecha_inicio = (hoy - timedelta(days=29)).isoformat()


    # =====================================================
    # QUERYSET BASE
    # =====================================================

    registros = PlantillaGenerada.objects.all()


    # =====================================================
    # APLICAR FILTROS
    # =====================================================

    if fecha_inicio:
        registros = registros.filter(
            fecha__date__gte=fecha_inicio
        )

    if fecha_fin:
        registros = registros.filter(
            fecha__date__lte=fecha_fin
        )

    if usuario:
        registros = registros.filter(
            usuario__username=usuario
        )

    if resultado:
        if resultado == "RECHAZADOS":

            registros = registros.exclude(
                resultado__in=[
                    "PROCEDE",
                    "NO PROCEDE"
                ]
            )

        else:

            registros = registros.filter(
                resultado=resultado
            )

    if distribuidor:
        registros = registros.filter(
            distribuidor=distribuidor
        )

    if busqueda:
        registros = registros.filter(
            Q(gestion__icontains=busqueda)
            | Q(cedula__icontains=busqueda)
            | Q(nombre_cliente__icontains=busqueda)
            | Q(distribuidor__icontains=busqueda)
            | Q(usuario__username__icontains=busqueda)
        )

    # =====================================================
    # CÉDULAS CON MÚLTIPLES GESTIONES
    # SOLO DEL PERÍODO FILTRADO
    # =====================================================

    duplicados = list(
        registros
        .exclude(cedula="")
        .exclude(cedula__isnull=True)
        .values("cedula", "nombre_cliente")
        .annotate(
            total=Count("id"),
            ultima_fecha=Max("fecha"),
            total_usuarios=Count("usuario", distinct=True),
        )
        .filter(total__gt=1)
        .order_by("-total")[:50]
    )


    # =====================================================
    # ESTADOS
    # =====================================================

    metricas = registros.aggregate(
        total=Count("id"),
        procede=Count("id", filter=Q(resultado="PROCEDE")),
        no_procede=Count("id", filter=Q(resultado="NO PROCEDE")),
        rechazados=Count(
            "id",
            filter=~Q(resultado__in=["PROCEDE", "NO PROCEDE"]),
        ),
    )
    registros_count = metricas["total"]
    procede = metricas["procede"]
    no_procede = metricas["no_procede"]
    rechazados = metricas["rechazados"]


    estados_json = [
        procede,
        no_procede,
        rechazados
    ]


    # =====================================================
    # GESTIONES POR DÍA
    # =====================================================

    gestiones_dia = (
        registros
        .annotate(
            dia=TruncDate("fecha")
        )
        .values("dia")
        .annotate(
            total=Count("id")
        )
        .order_by("dia")
    )


    datos_dia = [
        {
            "dia": (
                g["dia"].strftime("%d/%m/%Y")
                if g["dia"]
                else ""
            ),
            "total": g["total"]
        }
        for g in gestiones_dia
    ]

    # =====================================================
    # DUPLICIDAD POR DISTRIBUIDOR
    # =====================================================

    datos_duplicidad_distribuidor = list(
        registros
        .exclude(distribuidor="")
        .exclude(distribuidor__isnull=True)
        .exclude(cedula="")
        .exclude(cedula__isnull=True)
        .values("distribuidor")
        .annotate(
            gestiones_involucradas=Count("id"),
            cedulas_unicas=Count("cedula", distinct=True),
        )
        .annotate(
            gestiones_repetidas=(
                F("gestiones_involucradas") - F("cedulas_unicas")
            ),
        )
        .filter(gestiones_repetidas__gt=0)
        .order_by("-gestiones_repetidas", "distribuidor")[:10]
    )


    # =====================================================
    # USUARIOS CON MÁS GESTIONES
    # =====================================================

    usuarios = (
        registros
        .values("usuario__username")
        .annotate(
            total=Count("id")
        )
        .order_by("-total")[:10]
    )

    # =====================================================
    # ANÁLISIS POR USUARIO (seguro para plantilla)
    # =====================================================

    analisis_qs = (
        registros
        .values("usuario__id", "usuario__username")
        .annotate(
            gestiones=Count("id"),
            procede=Count("id", filter=Q(resultado="PROCEDE")),
            no_procede=Count("id", filter=Q(resultado="NO PROCEDE")),
            rechazados=Count("id", filter=~Q(resultado__in=["PROCEDE", "NO PROCEDE"]))
        )
        .order_by("-gestiones", "usuario__username")
    )

    analisis_usuarios = list(analisis_qs)

    for u in analisis_usuarios:
        total = u.get("gestiones") or 0
        u["tasa_procede"] = round((u.get("procede", 0) / total) * 100, 1) if total else 0
        u["tasa_no_procede"] = round((u.get("no_procede", 0) / total) * 100, 1) if total else 0
        u["tasa_rechazo"] = round((u.get("rechazados", 0) / total) * 100, 1) if total else 0

    promedio_gestiones_por_usuario = round(registros_count / len(analisis_usuarios), 1) if analisis_usuarios else 0

    for u in analisis_usuarios:
        total = u.get("gestiones") or 0
        u["muestra_suficiente"] = total >= 5
        u["diferencia_promedio"] = round(
            total - promedio_gestiones_por_usuario,
            1,
        )
        if not promedio_gestiones_por_usuario:
            u["nivel_carga"] = "Sin actividad"
            u["nivel_carga_clase"] = "secondary"
        elif total > promedio_gestiones_por_usuario * 1.25:
            u["nivel_carga"] = "Alta"
            u["nivel_carga_clase"] = "danger"
        elif total < promedio_gestiones_por_usuario * 0.75:
            u["nivel_carga"] = "Baja"
            u["nivel_carga_clase"] = "warning"
        else:
            u["nivel_carga"] = "Equilibrada"
            u["nivel_carga_clase"] = "success"

        if not u["muestra_suficiente"]:
            u["calidad_estado"] = "Muestra insuficiente"
            u["calidad_clase"] = "secondary"
        elif u["tasa_rechazo"] >= 30:
            u["calidad_estado"] = "Revisar"
            u["calidad_clase"] = "danger"
        else:
            u["calidad_estado"] = "Estable"
            u["calidad_clase"] = "success"

    tasa_procede_global = round(
        (procede / registros_count) * 100,
        1,
    ) if registros_count else 0
    tasa_rechazo_global = round(
        (rechazados / registros_count) * 100,
        1,
    ) if registros_count else 0

    usuario_mas_activo = analisis_usuarios[0] if analisis_usuarios else None

    usuario_mas_rechazos = None
    if analisis_usuarios:
        usuario_mas_rechazos = max(analisis_usuarios, key=lambda item: (item.get("tasa_rechazo", 0), item.get("rechazados", 0)))

    recomendaciones_analisis = construir_recomendaciones_analisis(
        analisis_usuarios=analisis_usuarios,
        promedio_gestiones_por_usuario=promedio_gestiones_por_usuario,
        total_gestiones=registros_count,
        total_duplicados=len(duplicados),
    )


    # =====================================================
    # DISTRIBUIDORES CON MÁS GESTIONES
    # =====================================================

    distribuidores = list(
        registros
        .exclude(distribuidor="")
        .values("distribuidor")
        .annotate(
            total=Count("id")
        )
        .order_by("-total")[:10]
    )


    # =====================================================
    # PLANTILLAS MÁS UTILIZADAS
    # =====================================================

    plantillas = (
        registros
        .values("nombre_plantilla")
        .annotate(
            total=Count("id")
        )
        .order_by("-total")[:10]
    )


    # =====================================================
    # INCONSISTENCIAS
    # POR AHORA DESACTIVADO
    # =====================================================

    inconsistencias = []


    # =====================================================
    # KPIs
    # =====================================================

    total_gestiones = registros_count

    total_usuarios = (
        registros
        .values("usuario")
        .distinct()
        .count()
    )

    total_distribuidores = (
        registros
        .exclude(distribuidor="")
        .values("distribuidor")
        .distinct()
        .count()
    )

    total_plantillas = (
        registros
        .values("nombre_plantilla")
        .distinct()
        .count()
    )

    lista_usuarios = User.objects.order_by("username")

    lista_distribuidores = (
        PlantillaGenerada.objects
        .exclude(distribuidor="")
        .exclude(distribuidor__isnull=True)
        .values_list("distribuidor", flat=True)
        .distinct()
        .order_by("distribuidor")
    )


    gestiones_ordenadas = (
        registros
        .select_related("usuario")
        .order_by("-fecha")
    )
    limites_validos = (25, 50, 100)
    try:
        limite_gestiones = int(request.GET.get("limite", 25))
    except (TypeError, ValueError):
        limite_gestiones = 25
    if limite_gestiones not in limites_validos:
        limite_gestiones = 25

    paginador_gestiones = Paginator(
        gestiones_ordenadas,
        limite_gestiones,
    )
    ultimas_gestiones = paginador_gestiones.get_page(
        request.GET.get("page")
    )
    parametros_paginacion = request.GET.copy()
    parametros_paginacion.pop("page", None)
    query_paginacion = parametros_paginacion.urlencode()
    parametros_analista = request.GET.copy()
    parametros_analista.pop("usuario", None)
    parametros_analista.pop("page", None)
    query_analista = parametros_analista.urlencode()

    # =====================================================
    # CENTRO DE ALERTAS
    # =====================================================

    alertas_sistema = []

    for item in duplicados:
        prioridad = "CRITICA" if item["total"] >= 5 else "ADVERTENCIA"
        alertas_sistema.append({
            "prioridad": prioridad,
            "tipo": "Gestiones duplicadas",
            "titulo": item["cedula"],
            "detalle": (
                f'{item["nombre_cliente"] or "Sin nombre"}: '
                f'{item["total"]} gestiones realizadas por '
                f'{item["total_usuarios"]} analista(s).'
            ),
            "cantidad": item["total"],
            "fecha": item["ultima_fecha"],
            "cedula": item["cedula"],
        })

    for item in analisis_usuarios:
        gestiones = item.get("gestiones") or 0
        tasa = item.get("tasa_rechazo") or 0
        if gestiones >= 5 and tasa >= 30:
            alertas_sistema.append({
                "prioridad": "CRITICA" if tasa >= 50 else "ADVERTENCIA",
                "tipo": "Tasa de rechazo",
                "titulo": item.get("usuario__username") or "Sin usuario",
                "detalle": (
                    f'{item.get("rechazados", 0)} rechazos de '
                    f'{gestiones} gestiones ({tasa}%).'
                ),
                "cantidad": item.get("rechazados", 0),
                "fecha": None,
                "cedula": "",
            })

    distribuidores_alerta = (
        registros
        .exclude(distribuidor="")
        .values("distribuidor")
        .annotate(
            total=Count("id"),
            rechazos=Count(
                "id",
                filter=~Q(resultado__in=["PROCEDE", "NO PROCEDE"]),
            ),
        )
        .filter(total__gte=5, rechazos__gt=0)
    )
    for item in distribuidores_alerta:
        tasa = round((item["rechazos"] / item["total"]) * 100, 1)
        if tasa >= 30:
            alertas_sistema.append({
                "prioridad": "CRITICA" if tasa >= 50 else "ADVERTENCIA",
                "tipo": "Distribuidor con rechazos",
                "titulo": item["distribuidor"],
                "detalle": (
                    f'{item["rechazos"]} rechazos de '
                    f'{item["total"]} gestiones ({tasa}%).'
                ),
                "cantidad": item["rechazos"],
                "fecha": None,
                "cedula": "",
            })

    orden_prioridad = {"CRITICA": 0, "ADVERTENCIA": 1, "INFORMATIVA": 2}
    alertas_sistema.sort(
        key=lambda alerta: (
            orden_prioridad.get(alerta["prioridad"], 9),
            -(alerta["fecha"].timestamp() if alerta["fecha"] else 0),
            -alerta["cantidad"],
        )
    )
    resumen_alertas = {
        "total": len(alertas_sistema),
        "criticas": sum(
            alerta["prioridad"] == "CRITICA"
            for alerta in alertas_sistema
        ),
        "advertencias": sum(
            alerta["prioridad"] == "ADVERTENCIA"
            for alerta in alertas_sistema
        ),
    }



    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        "reportes.html",
        {
            "duplicados": duplicados,
            "usuarios": usuarios,
            "analisis_usuarios": analisis_usuarios,
            "recomendaciones_analisis": recomendaciones_analisis,
            "promedio_gestiones_por_usuario": promedio_gestiones_por_usuario,
            "tasa_procede_global": tasa_procede_global,
            "tasa_rechazo_global": tasa_rechazo_global,
            "usuario_mas_activo": usuario_mas_activo,
            "distribuidores": distribuidores,
            "plantillas": plantillas,
            "inconsistencias": inconsistencias,
            "alertas_sistema": alertas_sistema,
            "resumen_alertas": resumen_alertas,

            "estados": estados_json,

            "procede": procede,
            "no_procede": no_procede,
            "rechazados": rechazados,

            "gestiones_dia": datos_dia,
            "duplicidad_distribuidor": datos_duplicidad_distribuidor,
            "ultimas_gestiones": ultimas_gestiones,

            "total_gestiones": total_gestiones,
            "total_usuarios": total_usuarios,
            "total_distribuidores": total_distribuidores,
            "total_plantillas": total_plantillas,
            "busqueda": busqueda,
            "limite_gestiones": limite_gestiones,
            "query_paginacion": query_paginacion,
            "query_analista": query_analista,
            "lista_usuarios": lista_usuarios,
            "lista_distribuidores": lista_distribuidores,

            # Mantener filtros
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "hora_inicio": hora_inicio,
            "hora_fin": hora_fin,
            "usuario_seleccionado": usuario,
            "resultado_seleccionado": resultado,
            "distribuidor_seleccionado": distribuidor,
            "periodo_seleccionado": periodo,
        }
    )

   
